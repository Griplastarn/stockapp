import streamlit as st
import math
from datetime import datetime
from openpyxl import Workbook, load_workbook
from fpdf import FPDF
import os
import base64
import pandas as pd

EXCEL_FIL = "vedlogg.xlsx"
PDF_FIL = "vedrapport.pdf"

def spara_till_excel(längd, diameter, volym, fast, travad):
    if os.path.exists(EXCEL_FIL):
        wb = load_workbook(EXCEL_FIL)
        ws = wb.active
        if ws["A" + str(ws.max_row)].value == "SUMMA":
            ws.delete_rows(ws.max_row)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Datum", "Längd (m)", "Diameter (cm)", "m³/stock", "m³fub", "m³s"])

    datum = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([datum, längd, diameter, round(volym, 3), round(fast, 3), round(travad, 3)])

    total_fub = sum(row[4] for row in ws.iter_rows(min_row=2, values_only=True) if isinstance(row[4], (int, float)))
    total_s = sum(row[5] for row in ws.iter_rows(min_row=2, values_only=True) if isinstance(row[5], (int, float)))
    ws.append(["SUMMA", "", "", "", round(total_fub, 3), round(total_s, 3)])

    wb.save(EXCEL_FIL)

def skapa_pdf():
    try:
        wb = load_workbook(EXCEL_FIL)
        ws = wb.active

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "Vedrapport", ln=True, align="C")

        pdf.set_font("Helvetica", "B", 12)
        pdf.ln(10)
        headers = ["Datum", "Längd", "Diameter", "m³/stock", "m³fub", "m³s"]
        for h in headers:
            pdf.cell(32, 8, h, border=1)
        pdf.ln()

        pdf.set_font("Helvetica", "", 11)
        for row in ws.iter_rows(min_row=2, values_only=True):
            pdf.set_font("Helvetica", "B", 11) if row[0] == "SUMMA" else pdf.set_font("Helvetica", "", 11)
            for cell in row[:6]:
                pdf.cell(32, 8, str(cell) if cell else "", border=1)
            pdf.ln()

        pdf.output(PDF_FIL)
        return True
    except Exception:
        return False

def skapa_download_länk(filnamn, knapptext):
    with open(filnamn, "rb") as f:
        data = f.read()
    b64 = base64.b64encode(data).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{filnamn}">{knapptext}</a>'
    return href

def rensa_data():
    if os.path.exists(EXCEL_FIL):
        os.remove(EXCEL_FIL)
    if os.path.exists(PDF_FIL):
        os.remove(PDF_FIL)
    st.success("All data har rensats!")

# --- Streamlit börjar här ---
st.set_page_config(page_title="Vedräknare", page_icon="🪵")
st.title("🪓 Vedräknare")

# Rensa-knapp
if st.button("🧹 Rensa allt"):
    rensa_data()

# Formulär för ny stock
with st.form("vedform", clear_on_submit=True):
    längd = st.number_input("Längd på stock (meter)", min_value=0.0, step=0.01, format="%.2f", key="längd")
    diameter = st.number_input("Diameter (cm)", min_value=0.0, step=0.1, format="%.1f", key="diameter")

    submitted = st.form_submit_button("Räkna och spara")

    if submitted and längd > 0 and diameter > 0:
        radie = diameter / 200
        volym = math.pi * radie**2 * längd
        fast = volym
        travad = volym * 1.6

        spara_till_excel(längd, diameter, volym, fast, travad)
        st.success(f"✅ Volym: {volym:.3f} m³\nFast mått: {fast:.3f} m³fub\nTravad: {travad:.3f} m³s\nLoggat i vedlogg.xlsx")
    elif submitted:
        st.warning("❗ Fyll i båda fälten med giltiga värden.")

# Export till PDF
if st.button("📄 Exportera till PDF"):
    if skapa_pdf():
        st.success("📄 PDF skapad: vedrapport.pdf")
    else:
        st.error("❌ Fel vid PDF-export")

# Nedladdningslänkar
if os.path.exists(EXCEL_FIL):
    st.markdown(skapa_download_länk(EXCEL_FIL, "📥 Ladda ner Excel-fil"), unsafe_allow_html=True)
if os.path.exists(PDF_FIL):
    st.markdown(skapa_download_länk(PDF_FIL, "📥 Ladda ner PDF-rapport"), unsafe_allow_html=True)

# Visa stocktabell
if os.path.exists(EXCEL_FIL):
    try:
        df = pd.read_excel(EXCEL_FIL)
        st.subheader("📊 Inmatade stockar")
        st.dataframe(df, use_container_width=True)
    except Exception as e:
        st.warning(f"Kunde inte läsa Excel-fil: {e}")
